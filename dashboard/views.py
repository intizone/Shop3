from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.core.exceptions import FieldError

from . import models
from .models import Product, Category, EnterProduct
from main import models
from . funcs import search_with_fields, pagenator_page

import openpyxl
from datetime import datetime
from openpyxl import load_workbook
import logging
logger = logging.getLogger(__name__)



def dashboard(request):
    return render(request, 'dashboard/dashboard.html')


# Category
def list_category(request):
    categorys = models.Category.objects.all()
    return render(request, 'category/list.html', {'categorys':categorys})

def create_category(request):
    if request.method == "POST":
        models.Category.objects.create(
            name = request.POST['name']
        )
        return redirect('dashboard:list_category')
    return render(request, 'category/create.html')

def detail_category(request, slug):
    category = models.Category.objects.get(slug=slug)
    if request.method == "POST":
        category.name = request.POST['name']
        category.save()
        return redirect('dashboard:list_category')
    return render(request, 'category/detail.html', {'category':category})


def delete_category(request, id):
    models.Category.objects.get(id=id).delete()
    return redirect('dashboard:list_category')

def categorys_count(request):
    categorys_count = models.Category.objects.all().count()
    return categorys_count


# Product
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

def product_list(request):
    products = models.Product.objects.all()
    categorys = models.Category.objects.all()
    
    # Filtering based on GET parameters
    name = request.GET.get('name')
    category_id = request.GET.get('category')
    quantity = request.GET.get('quantity')

    if name:
        products = products.filter(name__icontains=name)
    if category_id:
        products = products.filter(category__id=category_id)
    if quantity:
        products = products.filter(quantity=quantity)
    
    # Pagination
    paginator = Paginator(products, 10)  # Show 10 products per page
    page_number = request.GET.get('page')
    try:
        products = paginator.page(page_number)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        products = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        products = paginator.page(paginator.num_pages)

    context = {
        'products': products,
        'categorys': categorys,
    }
    return render(request, 'product/listing.html', context)

def product_detail(request, slug):
    product = models.Product.objects.get(slug=slug)
    categorys = models.Category.objects.all()
    enters = models.EnterProduct.objects.filter(product=product)
    context = {
        'product':product,
        'categorys':categorys,
        'enters':enters
    }
    return render(request, 'product/details.html', context)

def product_update(request, id):
    product = get_object_or_404(Product, id=id)
    categories = Category.objects.exclude(id=product.category_id)

    if request.method == "POST":
        product.name = request.POST.get('name', '')
        product.category = get_object_or_404(Category, id=request.POST.get('category_id', ''))
        product.description = request.POST.get('description', '')
        product.price = request.POST.get('price', '')
        product.currency = request.POST.get('currency', '')
        
        quantity_enter_notation = int(request.POST.get('quantity_enter_notation', 0))
        if request.POST.get('quantity_enter'):
            enter_quantity = int(request.POST.get('quantity_enter'))
        else:
            enter_quantity = 0
        
        # Create EnterProduct object
        enter_product = EnterProduct.objects.create(
            quantity=enter_quantity,
            quantity_enter_notation=quantity_enter_notation,
            product=product
        )
        
        
        # Toggle is_active for EnterProduct
        enter_product.is_active = not enter_product.is_active
        enter_product.save()
        
        return redirect('dashboard:product_list')
    
    context = {
        'product': product,
        'categories': categories
    }
    return render(request, 'product/update.html', context)

def product_create(request):
    categorys = models.Category.objects.all()
    context = {
        'categorys':categorys
    }
    if request.method == "POST":
        name = request.POST['name']
        description = request.POST['description']
        quantity = request.POST['quantity']
        price = request.POST['price']
        currency = request.POST['currency']
        baner_image = request.FILES['baner_image']
        category_id = request.POST['category_id']
        images = request.FILES.getlist('images')
        product = models.Product.objects.create(
            name=name,
            description = description,
            quantity=quantity,
            price=price,
            currency=currency,
            baner_image=baner_image,
            category_id=category_id
        )
        for image in images:
            models.ProductImage.objects.create(
                image=image,
                product=product
            )
        return redirect('dashboard:dashboard')
    return render(request, 'product/create.html', context)

def product_delete(request, id):
    models.Product.objects.get(id=id).delete()
    return redirect('dashboard:product_list')



# Auth
def sign_up(request):
    if request.method == "POST":
        User.objects.create_user(username=request.POST['username'],  
                                 password=request.POST['password'],
                                 is_staff=True,
                                 )
        return redirect('dashboard:dashboard')        
    return render(request, 'authentication/sign_up.html')

def sign_in(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']        
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('main:index')
        else:
            return HttpResponse("Username yoki paroldahatolik")
        
    return render(request, 'authentication/sign_in.html')

def sign_out(request):
    logout(request)
    return redirect('main:index')

@login_required
def user_update(request):
    if request.method == "POST":
        old_password = request.POST.get('old_password')
        new_password = request.POST.get('new_password')
        
        if old_password != new_password:
            user = request.user        
            if user.check_password(old_password):
                user.username = request.POST['username']
                user.set_password(new_password)
                user.save()                                
                update_session_auth_hash(request, user)     
                return redirect('dashboard:dashboard')       
        else:
            return HttpResponse("Parol o`zgarmagan")
        
    return render(request, 'authentication/update.html')


# enters
def create_enter(request):
    if request.method == 'POST':
        product_id = request.POST['product_id']
        quantity = int(request.POST['quantity'])
        models.EnterProduct.objects.create(
            product_id=product_id,
            quantity=quantity
        )
        return redirect('dashboard:list_enter')
    return render(request, 'dashboard/enter/create.html', {'products':models.Product.objects.all()})

def update_enter(request, id):
    if request.method == 'POST':
        quantity = int(request.POST['quantity'])
        enter = models.EnterProduct.objects.get(id=id)
        enter.quantity = quantity
        enter.save()
    return redirect('dashboard:list_enter')

def delete_enter(request, id):
    models.EnterProduct.objects.get(id=id).delete()
    return redirect('dashboard:list_enter')

def delete_enters(request, id):
    # Your delete logic here
    models.EnterProduct.objects.get(id=id).delete()
    
    # Redirect back to the previous page
    return redirect(request.META.get('HTTP_REFERER'))

def list_enter(request):
    result = search_with_fields(request)
    logger.debug(f"Result from search_with_fields: {result}")

    enters = models.EnterProduct.objects.all()  # Retrieve all objects initially
    try:
        if result:  # Apply filter only if there are any filter criteria
            enters = enters.filter(**result)
    except FieldError as err:
        key_to_delete = err.__doc__.split()[3][1:-1]
        if key_to_delete in result:
            del result[key_to_delete]
        enters = enters.filter(**result)

    
    logger.debug(f"Filtered queryset: {enters}")

    # Get the desired number of items per page, possibly from request.GET or a default value
    num_items_per_page = request.GET.get('num_items_per_page', 10)
    logger.debug(f"Number of items per page: {num_items_per_page}")

    paginated_enters = pagenator_page(enters, num_items_per_page, request)

    context = {'enters': paginated_enters}
    return render(request, 'dashboard/enter/list.html', context)



# excel
def excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, quantity, created_at = row
            print(f"Name: {name}, Quantity: {quantity}, Created At: {created_at}")
            try:
                product = models.Product.objects.get(name=name)
                enter_product = models.EnterProduct.objects.create(
                    product=product,  # Set the product attribute
                    quantity=quantity,
                    created_at=created_at,
                )
                print("yaratildi")
            except models.Product.DoesNotExist:
                print(f"Product with name '{name}' does not exist.")
            
        return redirect('dashboard:list_enter')
    return render(request, 'dashboard/enter/list.html')

def print_to_excel(request):
    products = models.Product.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Details"
    headers = ['№', 'Name', 'Description', 'Quantity', 'Currency', 'Category']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).value = header
    for row_idx, product in enumerate(products, start=2):
        ws.cell(row=row_idx, column=1).value = row_idx - 1  # Numbering
        ws.cell(row=row_idx, column=2).value = product.name
        ws.cell(row=row_idx, column=3).value = product.description
        ws.cell(row=row_idx, column=4).value = product.quantity
        ws.cell(row=row_idx, column=5).value = "So'm" if product.currency == 2 else "Dollar"
        ws.cell(row=row_idx, column=6).value = product.category.name
    column_widths = [5, 20, 40, 10, 10, 20]  # Adjust as needed
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=product_details.xlsx'
    wb.save(response)
    return response

def download_excel(request):
    enters = models.EnterProduct.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enter Details"
    headers = ['№', 'Maxsulot nomi', 'Soni', 'Sana']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).value = header
    for row_idx, enter in enumerate(enters, start=2):
        ws.cell(row=row_idx, column=1).value = row_idx - 1  # Numbering
        ws.cell(row=row_idx, column=2).value = enter.product.name if enter.product else enter.product_name
        ws.cell(row=row_idx, column=3).value = enter.quantity
        ws.cell(row=row_idx, column=4).value = enter.created_at.strftime('%Y-%m-%d %H:%M')
    column_widths = [5, 20, 10, 20]  # Adjust as needed
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=enter_details.xlsx'
    wb.save(response)
    return response

def download_excels(request, id):
    
    
    product = get_object_or_404(models.Product, id=id)
    
    enters = models.EnterProduct.objects.filter(product_id=id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enter Details"
    headers = ['№', 'Maxsulot nomi', 'Soni', 'Sana']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).value = header
    for row_idx, enter in enumerate(enters, start=2):
        ws.cell(row=row_idx, column=1).value = row_idx - 1  # Numbering
        ws.cell(row=row_idx, column=2).value = enter.product.name if enter.product else enter.product_name
        ws.cell(row=row_idx, column=3).value = enter.quantity
        ws.cell(row=row_idx, column=4).value = enter.created_at.strftime('%Y-%m-%d %H:%M')
    column_widths = [5, 20, 10, 20]  # Adjust as needed
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=enter_details.xlsx'
    wb.save(response)
    return response


# Search from page

def search_view(request):
    query = request.GET.get('q')
    if query:
        # Filter categories based on the search query
        categorys = Category.objects.filter(name__icontains=query)
    else:
        # If no query provided, return all categories
        categorys = Category.objects.all()
    return render(request, 'category/list.html', {'categorys': categorys, 'query': query})
